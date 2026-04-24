import customtkinter  # This tells Python what 'customtkinter' is
import tkinter
import sys as _sys
for _stream in (_sys.stdout, _sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except Exception:
        pass
from extractor import run_extraction # Linking to your other file
import customtkinter as ctk
import pandas as pd
from tkinter import messagebox, filedialog
import os
from datetime import datetime
import pdfplumber
import re
from openpyxl.styles import Font, Alignment
from datetime import datetime
today = datetime.now().strftime("%d/%m/%Y")
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from openpyxl.styles import Alignment, Font, Border
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl.styles import Alignment, Font, Border, Side
import sys
import glob
import customtkinter as ctk
from PIL import Image
import os
import customtkinter as ctk
from PIL import Image
import shutil
from tkinter import filedialog
import pandas as pd
from tkinter import ttk
import subprocess
import subprocess
import os
import subprocess
import platform
def load_documents_view(self, base_directory):
    # CHANGE THIS to your actual scrollable frame variable
    target_frame = self.main_scrollable_frame 

    # Clear existing UI
    for widget in target_frame.winfo_children():
        widget.destroy()

    if not os.path.exists(base_directory):
        print(f"DEBUG: Cannot find path: {base_directory}")
        return

    print(f"DEBUG: Scanning main folder: {base_directory}")

    for item_name in sorted(os.listdir(base_directory)):
        if item_name.startswith('.') or item_name.startswith('~$'):
            continue

        item_path = os.path.join(base_directory, item_name)

        if os.path.isdir(item_path):
            print(f"DEBUG: Found Sub-folder -> {item_name}")
            
            # 1. Create the Folder Frame
            folder_frame = ctk.CTkFrame(target_frame, fg_color="transparent")
            folder_frame.pack(fill="x", pady=5, padx=5)
            ctk.CTkLabel(folder_frame, text=f"📁 {item_name}", font=("Arial", 14, "bold")).pack(anchor="w", padx=5)

            # 2. Scan inside the Sub-folder
            for sub_item in sorted(os.listdir(item_path)):
                if sub_item.startswith('.') or sub_item.startswith('~$'):
                    continue
                    
                sub_item_path = os.path.join(item_path, sub_item)

                if os.path.isfile(sub_item_path):
                    print(f"DEBUG:   Found File -> {sub_item}")
                    
                    # 3. Create File Frame INSIDE Folder Frame
                    file_frame = ctk.CTkFrame(folder_frame)
                    file_frame.pack(fill="x", pady=2, padx=(30, 5)) 
                    ctk.CTkLabel(file_frame, text=f"📄 {sub_item}").pack(side="left", padx=10, pady=5)



def refresh_files(self):
    for widget in self.file_container.winfo_children():
        widget.destroy()

    files = get_all_files("Document")

    for file in files:
        self.create_file_row(file["relative"], file["path"])
def open_file(path):
    if platform.system() == "Darwin":  # macOS
        subprocess.call(["open", path])
    elif platform.system() == "Windows":
        os.startfile(path)
    else:  # Linux
        subprocess.call(["xdg-open", path])
def get_all_files(folder_path):
    all_files = []

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            full_path = os.path.join(root, file)

            # Optional: get relative path for display
            relative_path = os.path.relpath(full_path, folder_path)

            all_files.append({
                "name": file,
                "path": full_path,
                "relative": relative_path
            })

    return all_files

# Hàm quan trọng: Giúp file .exe xác định đúng thư mục đang đứng
def get_base_path():
    if getattr(sys, 'frozen', False):
        # Nếu là file .exe, lấy đường dẫn thư mục chứa file .exe
        return os.path.dirname(sys.executable)
    # Nếu đang chạy code .py trong VS Code
    return os.path.dirname(os.path.abspath(__file__))

def auto_update_schedule():
    # 1. Lấy ngày hiện tại
    today = datetime.now()
    date_str = today.strftime('%Y-%m-%d')
    vn_date = today.strftime('%d/%m/%Y')

    base_path = get_base_path()
    
    # 2. Tự động tìm file có tên chứa "schedule" trong thư mục
    search_pattern = os.path.join(base_path, "schedule*.*")
    files = glob.glob(search_pattern)
    
    if not files:
        print(f"[auto_update_schedule] Không tìm thấy file 'schedule' nào trong: {base_path}")
        return

    file_path = files[0]

    try:
        # 3. Đọc dữ liệu
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, skiprows=3)
        else:
            df = pd.read_excel(file_path, skiprows=3)

        df.columns = [str(col).replace('\n', ' ').strip() for col in df.columns]
        slot_col = [c for c in df.columns if 'Cặp' in c and 'tiết' in c][0]
        
        # Tìm cột ngày hôm nay
        date_column = next((c for c in df.columns if c.startswith(date_str)), None)

        if not date_column:
            print(f"[auto_update_schedule] Hôm nay ({vn_date}) không có lịch dạy trong file nguồn.")
            return

        # 4. Xử lý gộp ô và lọc dữ liệu
        df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill()
        df['MÔN HỌC'] = df['MÔN HỌC'].ffill()

        wb = Workbook()
        ws = wb.active
        
        # Định dạng style
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Tiêu đề và Header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"KẾ HOẠCH GIẢNG DẠY NGÀY {today.day} THÁNG {today.month} NĂM {today.year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align

        headers = ["Họ và tên", "môn học", "1 - 2", "3 - 4", "5 - 6", "7 - 8"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = center_align

        # 5. Ghi dữ liệu giáo viên
        current_row = 5
        exclude = ["THỐNG KÊ", "QUÂN SỰ", "QUỐC TẾ", "CÔNG AN", "TỔNG", "CỘNG", "SÁNG", "CHIỀU"]
        
        for teacher in df['HỌ VÀ TÊN'].dropna().unique():
            if any(key in str(teacher).upper() for key in exclude) or len(str(teacher)) < 2:
                continue

            teacher_df = df[df['HỌ VÀ TÊN'] == teacher]
            is_first = True
            
            for subject in teacher_df['MÔN HỌC'].unique():
                if any(key in str(subject).upper() for key in exclude): continue

                sub_df = teacher_df[teacher_df['MÔN HỌC'] == subject]
                slots = {"1 - 2": "", "3 - 4": "", "5 - 6": "", "7 - 8": ""}
                
                for _, row in sub_df.iterrows():
                    s = str(row[slot_col]).strip()
                    if "7 - 9" in s: s = "7 - 8"
                    if s in slots: slots[s] = row[date_column]

                row_vals = [teacher if is_first else "", subject, slots["1 - 2"], slots["3 - 4"], slots["5 - 6"], slots["7 - 8"]]
                for idx, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=current_row, column=idx, value=val)
                    cell.border = thin_border
                    cell.alignment = center_align
                is_first = False
                current_row += 1
            current_row += 1

        # Tự động căn chỉnh cột
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        for col in ['C','D','E','F']: ws.column_dimensions[col].width = 18

        # 6. Lưu file cùng thư mục với EXE
        output_path = os.path.join(base_path, f"KeHoach_Ngay_{date_str}.xlsx")
        wb.save(output_path)
        print(f"[auto_update_schedule] Đã cập nhật lịch ngày {vn_date}. File: {output_path}")

    except Exception as e:
        print(f"[auto_update_schedule] Lỗi hệ thống: {e}")

if __name__ == "__main__":
    auto_update_schedule()
# another
WEEKDAY_MAP = {
    0: "H",
    1: "B",  # Tuesday
    2: "T",  # Wednesday
    3: "N",  # Thursday
    4: "S",  # Friday
    5: "By", # Saturday
    6: "CN"  
}
TIME_WINDOWS = {
    "1-2": ("06:45", "08:15"),
    "3-4": ("08:25", "09:55"),
    "5-6": ("10:05", "11:25"),
    "7-8": ("13:45", "15:05")
}

def check_teaching_status(period_key):
    """Returns '1-2' if teaching, else 'He is out of class'"""
    now = datetime.now().time()
    clean_key = period_key.replace(" ", "") # Handles "1 - 2"
    
    if clean_key in TIME_WINDOWS:
        start_str, end_str = TIME_WINDOWS[clean_key]
        start = datetime.strptime(start_str, "%H:%M").time()
        end = datetime.strptime(end_str, "%H:%M").time()
        
        if start <= now <= end:
            return f"Teaching: {period_key}"
            
    return "He is out of class"

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")
COLORS = {
    "bg": "#F5F7F9",          # Nền chính (Xám nhẹ)
    "sidebar": "#FFFFFF",      # Sidebar (Trắng)
    "card": "#FFFFFF",         # Thẻ nội dung (Trắng)
    "accent": "#2563EB",       # Xanh dương chủ đạo
    "text": "#1E293B",         # Chữ chính (Đen xanh)
    "text_dim": "#64748B",     # Chữ phụ (Xám)
    "hover": "#F1F5F9",        # Màu khi di chuột qua
    "border": "#E2E8F0",       # Màu viền mảnh
    "success": "#10B981",      # Xanh lá (Dùng cho trạng thái sẵn sàng)
    "warning": "#F59E0B",      # Vàng cam
    "error": "#EF4444",        # Đỏ
    "purple": "#8B5CF6",       # Tím (Dự phòng cho các nút cũ)
    "orange": "#F97316"        # Cam (Dự phòng cho các nút cũ)
}#E2E8F0"        # Đường kẻ chia cắt


class DocumentFrame(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        
        # Configure grid
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Create Treeview
        self.tree = ttk.Treeview(self, columns=("filename"), show="headings")
        self.tree.heading("filename", text="Tên tài liệu (Double click to open)")
        self.tree.column("filename", anchor="w", width=400)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        # Bind double click
        self.tree.bind("<Double-1>", self.open_file)

        self.refresh_list()

    def refresh_list(self):
        path = "document"
        if not os.path.exists(path):
            os.makedirs(path)
            
        for file in os.listdir(path):
            if not file.startswith('.'): # Ignore hidden mac files
                self.tree.insert("", "end", values=(file,))

    def open_file(self, event):
        selected_item = self.tree.selection()[0]
        name = self.tree.item(selected_item, "values")[0]
        full_path = os.path.join("document", name)
        
        # Mac specific command to open Word/PDF
        subprocess.call(["open", full_path])
class TeacherCard(ctk.CTkFrame):
    def __init__(self, master, name, period, detail):
        super().__init__(master)
        
        self.period = period
        
        # Teacher Name & Detail (e.g., "td+ b6,7/c2")
        self.info_label = ctk.CTkLabel(self, text=f"{name} ({detail})", font=("Arial", 13))
        self.info_label.pack(side="left", padx=10)
        
        # Status Notification Label
        self.status_label = ctk.CTkLabel(self, text="", font=("Arial", 12, "bold"))
        self.status_label.pack(side="right", padx=10)
     
        self.update_status()  
class TeacherDetailWindow(ctk.CTkToplevel):

    def __init__(self, parent, data):
        super().__init__(parent)
        # Lấy tên giảng viên làm tiêu đề
        name = str(data.get('HỌ VÀ TÊN', 'CHI TIẾT')).upper()
        self.title(f"Thông tin: {name}")
        self.geometry("550x650")
        self.attributes("-topmost", True)  # Luôn hiện trên cùng
        self.configure(fg_color="#F1F5F9")

        # Container chính
        container = ctk.CTkFrame(self, fg_color="white", corner_radius=15)
        container.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(container, text="HỒ SƠ GIẢNG VIÊN", font=("Arial", 16, "bold"), text_color="#64748B").pack(pady=(15, 0))
        ctk.CTkLabel(container, text=name, font=("Arial", 24, "bold"), text_color="#1E40AF").pack(pady=(0, 20))

        # Vùng cuộn thông tin
        info_scroll = ctk.CTkScrollableFrame(container, fg_color="transparent")
        info_scroll.pack(fill="both", expand=True, padx=10)

        # Tự động quét qua tất cả các cột dữ liệu
        for key, value in data.items():
            # Bỏ qua các cột không cần thiết hoặc rác
            if "UNNAMED" in str(key).upper() or str(value).lower() == "nan":
                continue
                
            row_frame = ctk.CTkFrame(info_scroll, fg_color="#F8FAFC", corner_radius=8)
            row_frame.pack(fill="x", pady=3)

            # Tên trường thông tin (bên trái)
            ctk.CTkLabel(row_frame, text=str(key), font=("Arial", 12, "bold"), 
                         text_color="#475569", width=140, anchor="w").pack(side="left", padx=15, pady=10)
            
            # Giá trị thông tin (bên phải)
            ctk.CTkLabel(row_frame, text=str(value), font=("Arial", 13), 
                         text_color="#1E293B", wraplength=280, justify="left").pack(side="left", fill="x", expand=True, padx=5)

        ctk.CTkButton(container, text="ĐÓNG", fg_color="#1E293B", command=self.destroy).pack(pady=20)
   
class TeacherManagerPro(ctk.CTk):
    def clear_right_frame(self):
    # Ensure self.right_frame actually exists before trying to clear it
        if hasattr(self, 'right_frame'):
            for widget in self.right_frame.winfo_children():
                widget.destroy()
        else:
            print("Error: right_frame has not been initialized yet.")
    def hide_all_frames(self):
        self.mgmt_frame.pack_forget()
        self.plan_frame.pack_forget()
        self.document_frame.pack_forget()


    def build_tree(self, folder_path):
        def is_hidden(name):
            return name.startswith(".") or name.startswith("~$") or name.lower() == "thumbs.db"

        tree = {}

        for root, dirs, files in os.walk(folder_path):
            dirs[:] = [d for d in dirs if not is_hidden(d)]

            rel_path = os.path.relpath(root, folder_path)
            parts = rel_path.split(os.sep) if rel_path != "." else []

            current = tree
            for part in parts:
                current = current.setdefault(part, {})

            for file in files:
                if is_hidden(file):
                    continue
                current[file] = None

        return tree
    def render_documents(self):
        for child in self.document_scroll.winfo_children():
            child.destroy()

        folder = "Document"

        if not os.path.exists(folder):
            return

        tree = self.build_tree(folder)

        if not tree:
            ctk.CTkLabel(
                self.document_scroll,
                text="📂 Chưa có tài liệu nào",
                font=("Segoe UI", 16)
            ).pack(pady=40)
            return

        self.render_tree(self.document_scroll, tree)
    def render_tree(self, parent, tree, base_path="", level=0):
        for name, content in sorted(tree.items(), key=lambda x: (not isinstance(x[1], dict), x[0].lower())):
            full_path = os.path.join(base_path, name)
            if isinstance(content, dict):
                self._render_folder_node(parent, name, content, full_path, level)
            else:
                self._render_file_node(parent, name, full_path, level)

    def _render_folder_node(self, parent, name, content, full_path, level):
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="x", padx=0, pady=0, anchor="w")

        header = ctk.CTkFrame(container, fg_color=("#E2E8F0", "#1F2937"), corner_radius=6, height=36)
        header.pack(fill="x", padx=(10 + level * 22, 10), pady=2)
        header.pack_propagate(False)

        child_frame = ctk.CTkFrame(container, fg_color="transparent")
        state = {"open": False}

        arrow = ctk.CTkLabel(header, text="▸", font=("Segoe UI", 12, "bold"), width=18, cursor="hand2")
        arrow.pack(side="left", padx=(10, 0))
        icon = ctk.CTkLabel(header, text="📁", font=("Segoe UI", 13), cursor="hand2")
        icon.pack(side="left", padx=(2, 4))
        label = ctk.CTkLabel(header, text=name, font=("Segoe UI", 13, "bold"), cursor="hand2", anchor="w")
        label.pack(side="left", padx=2, fill="x", expand=True)

        def toggle(event=None):
            if state["open"]:
                child_frame.pack_forget()
                arrow.configure(text="▸")
                icon.configure(text="📁")
                state["open"] = False
            else:
                child_frame.pack(fill="x", anchor="w")
                arrow.configure(text="▾")
                icon.configure(text="📂")
                state["open"] = True

        for w in (header, arrow, icon, label):
            w.bind("<Button-1>", toggle)

        self.render_tree(child_frame, content, full_path, level + 1)

    def _render_file_node(self, parent, name, full_path, level):
        normal_color = ("#F8FAFC", "#111827")
        hover_color = ("#DBEAFE", "#1E3A8A")

        row = ctk.CTkFrame(parent, height=34, corner_radius=6, fg_color=normal_color)
        row.pack(fill="x", padx=(30 + level * 22, 10), pady=2)
        row.pack_propagate(False)

        icon = ctk.CTkLabel(row, text="📄", font=("Segoe UI", 13), cursor="hand2")
        icon.pack(side="left", padx=(10, 4))
        label = ctk.CTkLabel(row, text=name, font=("Segoe UI", 12), cursor="hand2", anchor="w")
        label.pack(side="left", padx=0, fill="x", expand=True)

        def on_click(event=None):
            self.open_document(full_path)

        def on_enter(event=None):
            row.configure(fg_color=hover_color)

        def on_leave(event=None):
            row.configure(fg_color=normal_color)

        for w in (row, icon, label):
            w.bind("<Button-1>", on_click)
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)

    def create_teacher_card(self, row):
        """Hàm phụ tạo từng dòng giảng viên"""
        card = ctk.CTkFrame(self.mgmt_scroll, fg_color="white", height=55, corner_radius=10, 
                            border_width=1, border_color="#E2E8F0")
        card.pack(fill="x", pady=2, padx=(20, 10))
        card.pack_propagate(False)

        # Hiển thị tên
        name_label = ctk.CTkLabel(card, text=row.get('HỌ VÀ TÊN', 'N/A'), font=("Arial", 14, "bold"))
        name_label.pack(side="left", padx=20)
        
        # Hiển thị cấp bậc (nếu có)
        rank = row.get('CẤP BẬC', '')
        if rank and str(rank) != "nan":
            ctk.CTkLabel(card, text=f"({rank})", font=("Arial", 12), text_color="#64748B").pack(side="left")

        # Nút bấm xem chi tiết
        # Truyền toàn bộ dữ liệu của dòng (row) vào cửa sổ mới
        btn = ctk.CTkButton(card, text="XEM CHI TIẾT", width=100, height=32, 
                            fg_color="#2563EB", hover_color="#1D4ED8",
                            command=lambda r=row.to_dict(): TeacherDetailWindow(self, r))
        btn.pack(side="right", padx=15)
   
    def show_document_frame(self):
        self.hide_all_frames()
        self.document_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_document)
        self.render_documents()
  
    def start_live_sync(self):
        if self.mgmt_data and self.plan_path: # Ensure files are linked
            # Update the data in memory
           
            # Re-render the UI
            self.render_mgmt()
            
        # Refresh every 60 seconds to keep "Real Time"
        self.after(60000, self.start_live_sync)
    def process_military_plan_with_calendar(file_path):
        teaching_data = []
        
        # Get current date info
        now = datetime.now()
        today_num = str(now.day).zfill(2) # "31"
        today_char = WEEKDAY_MAP[now.weekday()] # "B" (for Tuesday, March 31)

        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    table = page.extract_table()
                    if not table or len(table) < 2: continue
                    
                    # Row 0 is "Ngày" (01, 02, 03...)
                    # Row 1 is "Thứ" (T, N, S...)
                    days_row = table[0]
                    weekdays_row = table[1]
                    
                    target_col = None
                    for idx in range(len(days_row)):
                        # Check if column matches today's Date AND today's Weekday letter
                        if (days_row[idx] == today_num and 
                            weekdays_row[idx] == today_char):
                            target_col = idx
                            break
                    
                    if target_col is None: continue # Day not found on this page

                    # Process teachers in the rows below
                    for row in table[2:]:
                        # row[2] = Name, row[4] = Period (Tiết)
                        name = " ".join(str(row[2]).split()) if row[2] else None
                        period = str(row[4]).replace(" ", "") if row[4] else None
                        activity = row[target_col] # What is in today's column

                        if name and period and activity and activity.strip():
                            status = check_teaching_status(period)
                            teaching_data.append({
                                "teacher": name,
                                "period": period,
                                "detail": activity.strip(),
                                "notification": status
                            })
        except Exception as e:
            print(f"Error parsing PDF calendar: {e}")
        
        return teaching_data
    def sync_with_military_plan(self):
        # This automatically fetches today's specific assignments from the PDF[cite: 1]
        
        self.render_plan()
        # Check again every 5 minutes to see if a teacher has started a new slot
        self.after(300000, self.sync_with_military_plan)

    
    def __init__(self):
        super().__init__()
        self.title("TSQ Teacher Manager Pro")
        self.geometry("1300x850")
        self.configure(fg_color=COLORS["bg"])

        # 1. Khởi tạo dữ liệu
        self.mgmt_data = []
        self.plan_data = []
        
        # 2. Xây dựng Layout (Quan trọng: Tạo frame trước khi setup UI)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.setup_sidebar()
        
        self.main_container = ctk.CTkFrame(self, fg_color="transparent")
        self.main_container.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        
        self.mgmt_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.plan_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.document_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        # 3. Setup các thành phần giao diện
        self.setup_mgmt_ui()
        self.setup_plan_ui()
        self.setup_document_ui()
        # 4. Chạy các tính năng tự động
        self.show_mgmt_frame() # Hiện tab quản lý trước
        self.update_time()
        
        self.check_realtime_status() # Tự quét file kế hoạch ngày
        self.teacher_db = [] # Khởi tạo danh sách trống trước
        self.after(100, self.auto_load_mgmt_file)
    
       
    def setup_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=260, corner_radius=0, fg_color=COLORS["sidebar"])
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        
        ctk.CTkLabel(self.sidebar, text="TSQ QLGV", font=("Arial", 22, "bold"), text_color=COLORS["accent"]).pack(pady=40)
        
        self.btn_mgmt = self.create_nav_btn("Quản lý chung", self.show_mgmt_frame)
        self.btn_plan = self.create_nav_btn("Kế hoạch giảng", self.show_plan_frame)
        self.btn_document = self.create_nav_btn("Tài liệu môn học", self.show_document_frame) 
        
        self.lbl_time = ctk.CTkLabel(self.sidebar, text="", font=("Arial", 13), text_color=COLORS["text_dim"])
        self.lbl_time.pack(side="bottom", pady=30)

    def create_nav_btn(self, text, cmd):
        btn = ctk.CTkButton(self.sidebar, text=text, font=("Arial", 14, "bold"), height=52,
                            fg_color="transparent", text_color=COLORS["text"],
                            anchor="w", hover_color=COLORS["hover"], command=cmd)
        btn.pack(pady=5, padx=20, fill="x")
        return btn

    def set_active_nav(self, active_btn):
        for b in (self.btn_mgmt, self.btn_plan, self.btn_document):
            if b is active_btn:
                b.configure(fg_color=COLORS["accent"], text_color="white",
                            hover_color="#1D4ED8")
            else:
                b.configure(fg_color="transparent", text_color=COLORS["text"],
                            hover_color=COLORS["hover"])
    def load_excel_smart(self, path, check_cols):
        try:
            raw = pd.read_excel(path, header=None)
            header_row = None
            for i, row in raw.iterrows():
                row_vals = [str(x).upper() for x in row.values]
                if any("HỌ VÀ TÊN" in str(val) for val in row_vals):
                    header_row = i
                    break
            
            if header_row is None: return None
            df = pd.read_excel(path, skiprows=header_row)
            df.columns = [str(c).strip() for c in df.columns]
            cols_joined = " ".join(df.columns).upper()
            return df.to_dict('records') if any(col.upper() in cols_joined for col in check_cols) else None
        except: return None
        
    def setup_document_ui(self):
        header = ctk.CTkFrame(self.document_frame, fg_color="transparent")
        header.pack(fill="x", pady=10)

        ctk.CTkLabel(header, text="Tài liệu môn học", font=("Arial", 24, "bold")).pack(side="left", padx=20)

        ctk.CTkButton(header, text="Làm mới", command=self.render_documents).pack(side="right", padx=20)

        self.document_scroll = ctk.CTkScrollableFrame(self.document_frame)
        self.document_scroll.pack(fill="both", expand=True, padx=10, pady=10)
    # --- TAB: QUẢN LÝ CHUNG ---
    def render_mgmt(self):
        for widget in self.mgmt_scroll.winfo_children():
            widget.destroy()
            
        if not self.mgmt_data:
            print("⚠️ Cảnh báo: mgmt_data đang trống, không có gì để vẽ.")
            return

        search = self.mgmt_search.get().lower()
        
        count = 0
        for row in self.mgmt_data:
            name = str(row.get('HỌ VÀ TÊN', '')).upper()
            
            # Lọc theo ô tìm kiếm
            if search and search not in name.lower():
                continue

            # Tạo Card cho giảng viên
            card = ctk.CTkFrame(self.mgmt_scroll, fg_color="white", height=50, corner_radius=8)
            card.pack(fill="x", pady=2, padx=10)
            card.pack_propagate(False)

            ctk.CTkLabel(card, text=name, font=("Arial", 13, "bold")).pack(side="left", padx=15)
            
            # Nút Chi tiết - Gắn dữ liệu 'row' vào Popup
            btn = ctk.CTkButton(card, text="Chi tiết", width=80, height=30,
                                command=lambda r=row: TeacherDetailWindow(self, r))
            btn.pack(side="right", padx=10)
            count += 1
            
        print(f"Đã hiện thị {count} giảng viên lên màn hình.")
    def show_mgmt_frame(self):
        self.hide_all_frames()
        self.mgmt_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_mgmt)

    def setup_mgmt_ui(self):
        self.clear_right_frame()
        header = ctk.CTkFrame(self.mgmt_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(header, text="Thông tin giảng viên", font=("Arial", 28, "bold")).pack(side="left")

        self.mgmt_search = ctk.CTkEntry(self.mgmt_frame, placeholder_text="Tìm kiếm tên giảng viên hoặc khoa", height=40)
        self.mgmt_search.pack(fill="x", pady=10)
        self.mgmt_search.bind("<KeyRelease>", lambda e: self.render_mgmt())

        self.mgmt_scroll = ctk.CTkScrollableFrame(self.mgmt_frame, fg_color="transparent")
        self.mgmt_scroll.pack(fill="both", expand=True)
    def link_mgmt(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.mgmt_path = path
            self.refresh_mgmt()
    def refresh_mgmt(self):

        path = "danh sách k8.xlsx" # Hoặc dùng filedialog.askopenfilename()
        if not os.path.exists(path):
            print("Không tìm thấy file!")
            return

        try:
            # 1. Đọc file với engine openpyxl (quan trọng)
            # skiprows=2: Bỏ qua các dòng tiêu đề rỗng phía trên
            df = pd.read_excel(path, skiprows=2, engine='openpyxl', dtype=str)

            # 2. Chuẩn hóa tên cột: Xóa khoảng trắng và viết HOA toàn bộ
            df.columns = [str(c).strip().upper() for c in df.columns]

            # 3. Làm sạch dữ liệu: Xử lý gộp ô (ffill) và xóa dòng trống
            if 'HỌ VÀ TÊN' in df.columns:
                df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill() # Điền tên cho các ô bị gộp
                df = df.dropna(subset=['HỌ VÀ TÊN']) # Xóa dòng rác

            # 4. QUAN TRỌNG: Lưu vào biến self để các hàm khác có thể dùng
            self.mgmt_data = df.to_dict('records')
            
            # 5. Sau khi nhận được dữ liệu, gọi hàm vẽ giao diện ngay
            self.render_mgmt()
            print(f"Đã nhận {len(self.mgmt_data)} giảng viên từ file.")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể nhận dữ liệu từ file: {e}")
    def auto_load_mgmt_file(self):
        """Tự động tìm tiêu đề và nạp dữ liệu chính xác"""
        file_name = "danh sách k8.xlsx"
        if not os.path.exists(file_name):
            print(f"❌ Không tìm thấy file tại: {os.path.abspath(file_name)}")
            return

        try:
            # 1. Đọc nháp toàn bộ file (không dùng header) để tìm dòng tiêu đề
            raw_df = pd.read_excel(file_name, header=None, engine='openpyxl')
            
            header_row_index = None
            # Quét qua 20 dòng đầu tiên để tìm chữ "HỌ VÀ TÊN"
            for i, row in raw_df.head(20).iterrows():
                # Chuyển tất cả giá trị trong dòng thành chữ HOA, xóa khoảng trắng để so sánh
                row_values = [str(val).strip().upper() for val in row.values]
                if "HỌ VÀ TÊN" in row_values:
                    header_row_index = i
                    print(f"🎯 Đã tìm thấy tiêu đề 'HỌ VÀ TÊN' tại dòng thứ: {i + 1}")
                    break
            
            if header_row_index is None:
                print("❌ Vẫn không tìm thấy cột 'HỌ VÀ TÊN'.")
                print(f"Dữ liệu 5 dòng đầu đọc được:\n{raw_df.head(5)}")
                return

            # 2. Đọc lại file thật sự bắt đầu từ dòng tiêu đề đã tìm thấy
            df = pd.read_excel(file_name, skiprows=header_row_index, engine='openpyxl')
            
            # 3. Chuẩn hóa tên cột một lần nữa cho chắc chắn
            df.columns = [str(c).strip().upper() for c in df.columns]
            
            # 4. Làm sạch dữ liệu
            # Loại bỏ các cột "Unnamed" (cột thừa không có tên)
            df = df.loc[:, ~df.columns.str.contains('^UNNAMED')]
            
            # Điền đầy dữ liệu nếu có gộp ô (Merge Cells)
            if 'HỌ VÀ TÊN' in df.columns:
                df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill()
                df = df.dropna(subset=['HỌ VÀ TÊN']) # Xóa dòng hoàn toàn trống
                
                # Chuyển đổi sang danh sách Dictionary để dùng cho app
                self.mgmt_data = df.to_dict('records')
                
                # 5. Cập nhật giao diện
                self.render_mgmt()
                print(f"✅ Nạp thành công {len(self.mgmt_data)} giảng viên.")
            
        except Exception as e:
            print(f"❌ Lỗi xử lý: {e}")
    def process_mgmt_file(self, path):
        try:
            # 1. Đọc toàn bộ file không bỏ qua dòng nào để dò tìm
            raw_df = pd.read_excel(path, header=None, engine='openpyxl')
            
            header_row_index = None
            
            # 2. Vòng lặp tìm dòng chứa từ khóa "HỌ VÀ TÊN"
            for i, row in raw_df.iterrows():
                # Chuyển dòng thành danh sách chữ HOA để so sánh
                row_values = [str(val).strip().upper() for val in row.values]
                if "HỌ VÀ TÊN" in row_values:
                    header_row_index = i
                    break
            
            if header_row_index is None:
                print(f"❌ Không tìm thấy dòng nào chứa cột 'HỌ VÀ TÊN' trong file {path}")
                return

            # 3. Đọc lại file với đúng dòng tiêu đề đã tìm thấy
            df = pd.read_excel(path, skiprows=header_row_index, engine='openpyxl')
            
            # 4. Chuẩn hóa tên cột (Xóa khoảng trắng, viết HOA)
            df.columns = [str(c).strip().upper() for c in df.columns]
            
            # 5. Làm sạch dữ liệu rác
            # Điền đầy dữ liệu gộp ô (Merge cells)
            if 'HỌ VÀ TÊN' in df.columns:
                df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill()
                df = df.dropna(subset=['HỌ VÀ TÊN']) # Bỏ dòng trống hoàn toàn
                
                # Chuyển thành List Dict để dùng cho App
                self.mgmt_data = df.to_dict('records')
                
                # Vẽ lên màn hình
                self.render_mgmt()
                print(f"✅ Đã tìm thấy tiêu đề ở dòng {header_row_index + 1} và nạp thành công!")
            else:
                print("❌ Lỗi logic: Đã tìm thấy dòng tiêu đề nhưng không khớp cột.")

        except Exception as e:
            print(f"❌ Lỗi xử lý file: {e}")
   
#--------------------------
    def show_plan_frame(self):
        self.hide_all_frames()
        self.plan_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_plan)

    def render_plan(self):
        try:
            if not self.plan_scroll.winfo_exists(): return
            for child in self.plan_scroll.winfo_children():
                child.destroy()

            if not self.plan_data: return

            SUB_COLORS = {
                "BC": ("#E0F2FE", "#0369A1"), "ĐH": ("#DCFCE7", "#15803D"),
                "KB": ("#F3E8FF", "#7E22CE"), "ĐN": ("#FEF3C7", "#B45309"),
            }

            # --- HEADER ---
            header_f = ctk.CTkFrame(self.plan_scroll, fg_color="#F8FAFC", height=45, corner_radius=0)
            header_f.pack(fill="x", padx=5, pady=(0, 5))
            header_f.pack_propagate(False)
            
            COLS = [("Họ và tên", 0.02, 0.28), ("Môn", 0.30, 0.10), ("1-2", 0.42, 0.14), 
                    ("3-4", 0.57, 0.14), ("5-6", 0.72, 0.14), ("7-8", 0.87, 0.14)]
            for txt, rx, rw in COLS:
                ctk.CTkLabel(header_f, text=txt, font=("Arial", 12, "bold"), text_color="#64748B", anchor="w").place(relx=rx, rely=0.5, anchor="w", relwidth=rw)

            # --- DỮ LIỆU ---
            prev_name = ""
            group_frame = None # Frame chứa các dòng của cùng 1 người

            for i, row in enumerate(self.plan_data):
                def get_v(k):
                    v = str(row.get(k, "")).strip()
                    return "" if v.lower() == "nan" or v == "" else v

                full_name = get_v("Họ và tên")
                subject = get_v("môn học")
                if not full_name and not subject: continue

                # KIỂM TRA: Nếu có tên mới -> Tạo Group Frame mới và vẽ đường kẻ cho người cũ
                if full_name != "" and full_name != prev_name:
                    # Vẽ đường kẻ ngăn cách nếu đây không phải người đầu tiên
                    if i > 0:
                        ctk.CTkFrame(self.plan_scroll, fg_color="#CBD5E1", height=2).pack(fill="x", padx=5, pady=(2, 5))
                    
                    # Tạo Group Frame mới cho giảng viên này
                    group_frame = ctk.CTkFrame(self.plan_scroll, fg_color="white", corner_radius=0)
                    group_frame.pack(fill="x", padx=5)
                    prev_name = full_name
                    is_duplicate = False
                else:
                    is_duplicate = True

                # Tạo dòng bên trong group_frame
                row_f = ctk.CTkFrame(group_frame, fg_color="transparent", height=42, corner_radius=0)
                row_f.pack(fill="x")
                row_f.pack_propagate(False)

                # Vẽ Tên (Chỉ hiện ở dòng đầu của group)
                display_name = full_name if not is_duplicate else ""
                ctk.CTkLabel(row_f, text=display_name, font=("Arial", 13, "bold"), text_color="#1E293B", anchor="w").place(relx=0.02, rely=0.5, anchor="w", relwidth=0.28)
                
                # Vẽ Môn (Badge)
                if subject:
                    bg, fg = SUB_COLORS.get(subject.upper(), ("#F1F5F9", "#475569"))
                    badge = ctk.CTkFrame(row_f, fg_color=bg, corner_radius=6, height=22)
                    badge.place(relx=0.30, rely=0.5, anchor="w", relwidth=0.08)
                    ctk.CTkLabel(badge, text=subject, font=("Arial", 10, "bold"), text_color=fg).pack(expand=True)

                # Vẽ Lịch dạy
                times = ["1 - 2", "3 - 4", "5 - 6", "7 - 8"]
                for idx, t_col in enumerate(times):
                    val = get_v(t_col)
                    if val:
                        ctk.CTkLabel(row_f, text=val, font=("Arial", 11), text_color="#2563EB", anchor="w").place(relx=0.42 + (idx*0.15), rely=0.5, anchor="w", relwidth=0.14)

                # Vẽ đường kẻ mờ nội bộ (Nếu vẫn còn dòng tiếp theo của cùng 1 người)
                if i + 1 < len(self.plan_data):
                    next_name = get_v(self.plan_data[i+1].get("Họ và tên", ""))
                    if next_name == "" or next_name == prev_name:
                        ctk.CTkFrame(group_frame, fg_color="#F1F5F9", height=1).pack(fill="x", padx=10)

            self.update_idletasks()
        except Exception as e:
            print(f"❌ Lỗi: {e}")
    def setup_plan_ui(self):
        self.clear_right_frame()
        # Header
        header = ctk.CTkFrame(self.plan_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 20))
        
        ctk.CTkLabel(header, text="Kế hoạch giảng dạy trong ngày", 
                     font=("Arial", 26, "bold"), text_color=COLORS["text"]).pack(side="left")

        # Toolbar
        tool_bar = ctk.CTkFrame(self.plan_frame, fg_color="white", corner_radius=10, 
                                border_width=1, border_color=COLORS["border"])
        tool_bar.pack(fill="x", pady=10)

        self.status_indicator = ctk.CTkLabel(tool_bar, text="🟢 Hệ thống sẵn sàng", 
                                             font=("Arial", 12), text_color=COLORS["success"])
        self.status_indicator.pack(side="left", padx=20, pady=10)

        ctk.CTkButton(tool_bar, text="Làm mới dữ liệu", width=120, height=32,
                      fg_color=COLORS["accent"], hover_color="#1D4ED8",
                      font=("Arial", 11, "bold"),
                      command=self.check_realtime_status).pack(side="right", padx=15)

        # Khung cuộn hiển thị kế hoạch
        self.plan_scroll = ctk.CTkScrollableFrame(self.plan_frame, fg_color="transparent")
        self.plan_scroll.pack(fill="both", expand=True)
    def link_plan(self):
        """Hàm chọn file Excel từ máy tính"""
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.plan_path = path
            self.refresh_plan_data()
    def refresh_plan_data(self):
        if not self.plan_path:
            return

        try:
            df = pd.read_excel(self.plan_path, skiprows=3)
            df.columns = [str(col).strip() for col in df.columns]

            # Fill merged cells
            df['Họ và tên'] = df['Họ và tên'].ffill()

            slots = ["1 - 2", "3 - 4", "5 - 6", "7 - 8"]
            df = df.dropna(subset=['môn học'] + slots, how='all')

            # 🔥 GROUP BY TEACHER
            grouped = []

            for name, group in df.groupby('Họ và tên', sort=False):
                teacher = {
                    "name": name,
                    "subjects": [],
                    "rows": group.to_dict('records')
                }

                for _, r in group.iterrows():
                    subject = str(r.get('môn học', ''))
                    if subject != "nan" and subject not in teacher["subjects"]:
                        teacher["subjects"].append(subject)

                grouped.append(teacher)

            self.plan_data = df.to_dict('records')
            self.render_plan()

        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
    def open_document(self, file_name):
        path = os.path.join("Document", file_name)
        try:
            if os.name == "nt":
                os.startfile(path)
            else:
                subprocess.call(["open", path])
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
    def convert_excel_date(self, val):
        """Hàm phụ trợ để xử lý ngày tháng từ số Excel sang chuỗi dd/mm/yyyy"""
        try:
            if isinstance(val, (int, float)) and val > 1000:
                return pd.to_datetime(val, unit='D', origin='1899-12-30').strftime('%d/%m/%Y')
            return str(val) if str(val).lower() != 'nan' else ""
        except:
            return str(val)
    def update_time(self):
        self.lbl_time.configure(text=datetime.now().strftime("%H:%M:%S\n%A, %d/%m/%Y"))
        self.after(1000, self.update_time)   
    def load_monthly_plan(self, file_path):
        try:
            df = pd.read_excel(file_path)

            # Clear old content if reload
            for widget in self.tab_plan.winfo_children():
                widget.destroy()

            # Frame container
            frame = ctk.CTkFrame(self.tab_plan)
            frame.pack(fill="both", expand=True)

            # Create table
            tree = ttk.Treeview(frame)
            tree.pack(side="left", fill="both", expand=True)

            # Scrollbars
            scrollbar_y = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")

            scrollbar_x = ttk.Scrollbar(self.tab_plan, orient="horizontal", command=tree.xview)
            scrollbar_x.pack(fill="x")

            tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

            # Columns
            tree["columns"] = list(df.columns)
            tree["show"] = "headings"

            for col in df.columns:
                tree.heading(col, text=col)
                tree.column(col, anchor="center", width=120)

            # Rows
            for _, row in df.iterrows():
                tree.insert("", "end", values=list(row))

        except Exception as e:
            print("ERROR loading Excel:", e)   
    def load_documents(self, folder_path):
        frame = ctk.CTkFrame(self.tab_docs)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        files = [f for f in os.listdir(folder_path) if f.endswith(".pdf")]

        for file in files:
            btn = ctk.CTkButton(
                frame,
                text=file,
                anchor="w",
                command=lambda f=file: self.open_pdf(os.path.join(folder_path, f))
            )
            btn.pack(fill="x", pady=5)
    def check_realtime_status(self):
        import glob
        files = glob.glob("KeHoach_Ngay_*.xlsx")
        if not files: 
            print("❌ Không tìm thấy file Excel nào!")
            return

        latest_file = max(files, key=os.path.getctime)
        print(f"📂 Đang đọc file: {latest_file}")
        
        try:
            # Đọc từ dòng 4
            df = pd.read_excel(latest_file, skiprows=3)
            df.columns = [str(c).strip() for c in df.columns]
            
            # XỬ LÝ QUAN TRỌNG: Loại bỏ các dòng hoàn toàn trống
            # Chỉ giữ lại dòng có tên HOẶC có môn học
            df = df.dropna(subset=['Họ và tên', 'môn học'], how='all')
            
            self.plan_data = df.to_dict('records')
            
            # KIỂM TRA: In ra số lượng dòng Python đọc được
            print(f"✅ Đã nạp được {len(self.plan_data)} dòng dữ liệu.")
            
            if self.plan_scroll.winfo_exists():
                self.render_plan()
        except Exception as e:
            print(f"❌ Lỗi đọc file: {e}")


if __name__ == "__main__":
    app = TeacherManagerPro()
    app.mainloop()